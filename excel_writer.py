"""
excel_writer.py
Generates a 3-sheet colored Excel report:
  1. "Business Info"   — the data the user typed into the form
  2. "Comparison"       — CORRECT / INCORRECT (+ extracted value) / MISSING / N/A per field
  3. "Extracted Data"   — the raw value extracted from each live page per field

Supports mixed-source reports — all ALL_FIELDS as columns,
N/A shown in gray for fields not applicable to a given source.
"""

import io
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fields_config import ALL_FIELDS, SOURCE_FIELDS, VISUAL_FIELDS

# ── Color fills ───────────────────────────────────────────────────────────────
FILL_RED           = PatternFill("solid", fgColor="FF0000")
FILL_GRAY          = PatternFill("solid", fgColor="F2F2F2")
FILL_HEADER        = PatternFill("solid", fgColor="4472C4")
FILL_STATUS_HEADER = PatternFill("solid", fgColor="7030A0")

# ── Fonts ─────────────────────────────────────────────────────────────────────
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_RED    = Font(color="FFFFFF", bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_GRAY   = Font(color="888888", size=10)

# ── Alignment ─────────────────────────────────────────────────────────────────
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Non-wrapping variants — used on sheets where cell contents can be long
# free text (Description, Hours, Social Media Links, ...), so Excel never
# auto-expands the row height to fit every wrapped line.
ALIGN_CENTER_NOWRAP = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT_NOWRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# Fixed row height (points) used for data rows on sheets that disable wrap,
# so every row stays a uniform, compact size regardless of content length.
DATA_ROW_HEIGHT = 18

# ── Borders ───────────────────────────────────────────────────────────────────
THIN         = Side(style="thin", color="CCCCCC")
BORDER_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK_PURPLE = Side(style="medium", color="7030A0")
BORDER_STATUS_HEADER = Border(
    left=THICK_PURPLE, right=THICK_PURPLE,
    top=THICK_PURPLE,  bottom=THICK_PURPLE,
)

# Exact-match values that should always render as a red "attention" cell.
# NOTE: "INCORRECT" is also matched as a *prefix* (see _is_red_value) since
# the Comparison sheet now appends the extracted value, e.g.
# "INCORRECT /roofing contractor, Siding Contractor".
_RED_VALUES = {"INCORRECT", "MISSING", "SCRAPE ERROR"}


def _is_red_value(value) -> bool:
    if not isinstance(value, str):
        return False
    if value in _RED_VALUES:
        return True
    return value.startswith("INCORRECT")


def make_filename(business_name: str) -> str:
    """
    Convert the business name into a safe filename.
    e.g. "HAQQ Legal AI" -> "HAQQ_Legal_AI_listing_report.xlsx"
    Falls back to "listing_checker_report.xlsx" if name is blank.
    """
    name = (business_name or "").strip()
    if not name:
        return "listing_checker_report.xlsx"
    safe = re.sub(r"[^\w\s\-]", "", name)
    safe = re.sub(r"\s+", "_", safe).strip("_")
    return f"{safe}_listing_report.xlsx" if safe else "listing_checker_report.xlsx"


def _style_cell(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _style_value_cell(cell, value, nowrap=False):
    """Shared coloring rule used by both the Comparison and Extracted Data
    sheets: red for problem values, gray for N/A, plain otherwise.

    nowrap=True disables text wrapping (used on the Extracted Data sheet,
    where cell contents can be long free text) so Excel never auto-expands
    the row height to fit every wrapped line.
    """
    align_center = ALIGN_CENTER_NOWRAP if nowrap else ALIGN_CENTER
    align_left   = ALIGN_LEFT_NOWRAP if nowrap else ALIGN_LEFT

    if _is_red_value(value):
        _style_cell(cell, fill=FILL_RED, font=FONT_RED,
                    alignment=align_center, border=BORDER_THIN)
    elif value == "N/A":
        _style_cell(cell, fill=FILL_GRAY, font=FONT_GRAY,
                    alignment=align_center, border=BORDER_THIN)
    elif value == "CORRECT":
        _style_cell(cell, font=FONT_NORMAL, alignment=align_center, border=BORDER_THIN)
    else:
        _style_cell(cell, font=FONT_NORMAL, alignment=align_left, border=BORDER_THIN)


def _autosize_columns(ws, headers, explicit_widths=None):
    explicit_widths = explicit_widths or {}
    for col_idx, header in enumerate(headers, start=1):
        if header in explicit_widths:
            width = explicit_widths[header]
        else:
            width = max(len(header) + 4, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_header_row(ws, headers, status_col_header=None):
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        if header == status_col_header:
            _style_cell(cell,
                fill=FILL_STATUS_HEADER, font=FONT_HEADER,
                alignment=ALIGN_CENTER, border=BORDER_STATUS_HEADER)
        else:
            _style_cell(cell,
                fill=FILL_HEADER, font=FONT_HEADER,
                alignment=ALIGN_CENTER, border=BORDER_THIN)
    ws.row_dimensions[1].height = 30


# ── Sheet 1: Business Info ────────────────────────────────────────────────────

def _write_business_info_sheet(ws, user_data: dict):
    """
    Simple two-column reference sheet: the raw values the user typed into
    the form, in ALL_FIELDS order. This is the "ground truth" the other
    two sheets were checked against.
    """
    headers = ["Field", "Your Input"]
    _write_header_row(ws, headers)

    for field in ALL_FIELDS:
        value = (user_data or {}).get(field, "")
        # Visual fields are stored as "present"/"" sentinels in user_data
        # (see app.py) rather than free text — show something readable.
        if field in ("Logo", "Photos"):
            value = "Should be present" if value else "Not required"
        # Any field left blank on the form -> show "N/A" instead of an
        # empty cell, so it reads as "not provided" rather than "missing".
        if isinstance(value, str) and not value.strip():
            value = "N/A"
        ws.append([field, value])
        row_idx = ws.max_row
        _style_cell(ws.cell(row_idx, 1),
            font=Font(bold=True, size=10), alignment=ALIGN_LEFT, border=BORDER_THIN)
        if value == "N/A":
            _style_cell(ws.cell(row_idx, 2),
                fill=FILL_GRAY, font=FONT_GRAY, alignment=ALIGN_LEFT, border=BORDER_THIN)
        else:
            _style_cell(ws.cell(row_idx, 2),
                font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_THIN)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A2"


# ── Sheet 2: Comparison (CORRECT/INCORRECT/MISSING/N/A) ───────────────────────

def _extracted_value_for(field: str, extracted_row: dict):
    """Pull the raw extracted value for `field` out of an Extracted Data
    row dict, returning None if there's nothing usable to show."""
    if not extracted_row:
        return None
    raw = extracted_row.get(field)
    if raw in (None, ""):
        return None
    return str(raw)


def _write_comparison_sheet(ws, results: list, extracted_by_url: dict = None):
    """
    extracted_by_url: {live_link: extracted_row_dict} — used so that any
    field marked INCORRECT can show what was actually scraped, e.g.
    "INCORRECT /roofing contractor, Siding Contractor" instead of a bare
    "INCORRECT".
    """
    extracted_by_url = extracted_by_url or {}
    headers = ["Source", "Live Link", "Status"] + ALL_FIELDS
    _write_header_row(ws, headers, status_col_header="Status")

    for result in results or []:
        live_link = result.get("Live Link", "")
        extracted_row = extracted_by_url.get(live_link, {})

        row_values = [
            result.get("Source", ""),
            live_link,
            result.get("Status", ""),
        ]
        for field in ALL_FIELDS:
            status = result.get(field, "N/A")
            if status == "INCORRECT":
                extracted_val = _extracted_value_for(field, extracted_row)
                status = f"INCORRECT /{extracted_val}" if extracted_val else "INCORRECT"
            row_values.append(status)

        ws.append(row_values)
        row_idx = ws.max_row
        for col_idx, value in enumerate(row_values, start=1):
            _style_value_cell(ws.cell(row_idx, col_idx), value)

    _autosize_columns(ws, headers, {"Source": 22, "Live Link": 45, "Status": 14})
    ws.freeze_panes = "A2"


# ── Sheet 3: Extracted Data (raw scraped values) ──────────────────────────────

def _fields_allowed_for_source(source: str):
    """Mirrors data_extractor._normalize_extracted's rule: if the source is
    recognised, only its configured fields are tracked; unrecognised
    sources have every field tracked."""
    if source in SOURCE_FIELDS:
        return set(SOURCE_FIELDS[source])
    return None  # None = no restriction, every field allowed


def _write_extracted_sheet(ws, extracted_list: list, url_to_source: dict, user_data: dict = None):
    """
    user_data : the dict of values the user typed into the form. When a
    field was left blank on the form, that field is shown as "N/A" on
    this sheet too — mirroring the same "blank input -> N/A" rule the
    Comparison sheet already applies (see comparator.compare_row).
    Visual fields (Logo/Photos) are exempt, since they're graded on
    presence alone and have no user-typed value to be "blank".
    """
    user_data = user_data or {}
    headers = ["Source", "Live Link", "Status"] + ALL_FIELDS
    _write_header_row(ws, headers, status_col_header="Status")

    for row in extracted_list or []:
        url          = row.get("_url", "")
        source       = url_to_source.get(url, "unknown")
        scrape_error = row.get("_scrape_error")
        allowed      = _fields_allowed_for_source(source)

        row_values = [
            source,
            url,
            "SCRAPE ERROR" if scrape_error else "OK",
        ]
        for field in ALL_FIELDS:
            user_val = str(user_data.get(field, "")).strip()

            if allowed is not None and field not in allowed:
                value = "N/A"
            elif scrape_error:
                value = "SCRAPE ERROR"
            elif field not in VISUAL_FIELDS and not user_val:
                # User left this field blank on the form -> N/A, regardless
                # of whether the page happened to have a value.
                value = "N/A"
            else:
                raw = row.get(field, "")
                value = raw if raw not in (None, "") else "MISSING"
            row_values.append(value)

        ws.append(row_values)
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
        for col_idx, value in enumerate(row_values, start=1):
            _style_value_cell(ws.cell(row_idx, col_idx), value, nowrap=True)

    _autosize_columns(ws, headers, {"Source": 22, "Live Link": 45, "Status": 14})
    ws.freeze_panes = "A2"


# ── Public entry point ────────────────────────────────────────────────────────

def write_excel(results: list, extracted_list: list = None, user_data: dict = None) -> bytes:
    """
    Build the 3-sheet Excel report.

    results        : list of comparison dicts from comparator.compare_all()
                      (each has Source / Live Link / Status / <ALL_FIELDS> -> status)
    extracted_list : list of raw-extraction dicts from data_extractor.extract_batch()
                      (each has _url / <tracked fields> -> extracted value, or
                      _scrape_error on failure). Optional for backward compatibility.
    user_data      : the dict of values the user typed into the form. Optional.
    """
    wb = openpyxl.Workbook()

    ws_business = wb.active
    ws_business.title = "Business Info"
    _write_business_info_sheet(ws_business, user_data or {})

    extracted_by_url = {r.get("_url", ""): r for r in (extracted_list or [])}

    ws_comparison = wb.create_sheet("Comparison")
    _write_comparison_sheet(ws_comparison, results, extracted_by_url)

    ws_extracted = wb.create_sheet("Extracted Data")
    url_to_source = {r.get("Live Link", ""): r.get("Source", "unknown") for r in (results or [])}
    _write_extracted_sheet(ws_extracted, extracted_list or [], url_to_source, user_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
